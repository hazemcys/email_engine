import mailbox
import re
import pandas as pd
from collections import Counter
from tqdm import tqdm
import os, sys, argparse, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from classifier import run_ai_classification

from google import genai
import time, os


# Mapping column widths to make our Excel files look gorgeous natively!
_COL_W = {
    'sender_id':     38,
    'sender_email':  32,
    'sender_emails': 48,
    'title':         42,
    'body':          68,
    'phone':         14,
    'phones':        25,
    'report_phone':  14,
    'report_name':   38,
    'occurrences':   13,
    'category':      18,
}
_RIGHT = {'sender_id', 'title', 'body', 'report_name', 'sender_emails'}
_CENTER = {'phone', 'report_phone', 'occurrences', 'phones', 'category'}


def save_excel(records, filename):
    if not records:
        Workbook().save(filename)
        return

    def _clean(k, v):
        if isinstance(v, list):
            return ', '.join(str(x) for x in v)
        if not isinstance(v, str):
            return v
        # kill newlines so rows dont explode in height
        v = v.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
        if k == 'body' and len(v) > 700:
            return v[:700] + ' ...'
        return v

    clean = [{k: _clean(k, v) for k, v in r.items()} for r in records]
    headers = list(clean[0].keys())

    wb = Workbook()
    ws = wb.active

    hdr_fill  = PatternFill('solid', fgColor='1F4E79')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    hdr_align = Alignment(horizontal='center', vertical='center')

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill      = hdr_fill
        c.font      = hdr_font
        c.alignment = hdr_align
    ws.row_dimensions[1].height = 22

    fill_e    = PatternFill('solid', fgColor='EDF4FC')
    fill_o    = PatternFill('solid', fgColor='FFFFFF')
    data_font = Font(size=10, name='Calibri')

    for ri, record in enumerate(clean, 2):
        ws.row_dimensions[ri].height = 18
        bg = fill_e if ri % 2 == 0 else fill_o
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=record.get(h, ''))
            c.fill = bg
            c.font = data_font
            if h in _RIGHT:
                c.alignment = Alignment(horizontal='right', vertical='center')
            elif h in _CENTER:
                c.alignment  = Alignment(horizontal='center', vertical='center')
                c.number_format = '@'
            else:
                c.alignment = Alignment(horizontal='left', vertical='center')

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _COL_W.get(h, 22)

    ws.freeze_panes = 'A2'
    wb.save(filename)


# Translating Arabic digits! 
# We need this mapping because sometimes users type ٠١٢٣٤٥٦٧٨٩ instead of 0-9.
_AR = "٠١٢٣٤٥٦٧٨٩"
_EN = "0123456789"
_numtable = str.maketrans(_AR, _EN)

def fix_numbers(s):
    if not s: return ""
    return str(s).translate(_numtable)

# phone regex - tried lookahead before but caused issues on some mbox files
def pull_phones(txt):
    # strip KW country code first
    cleaned = re.sub(r'(\+965|00965|(?<!\d)965)(?=\d{8})', '', fix_numbers(txt))
    found = re.findall(r'\b\d{8}\b', cleaned)
    # deduplicate but keep them as list for now
    seen = []
    for p in found:
        if p not in seen:
            seen.append(p)
    return seen


# === THE LEGAL & SENSITIVE KEYWORDS ===
# We pull this right from the BRD specs! 
# Note: DO NOT sort this array! We put the longest compound words first on purpose 
# so 'قائد القوات' gets matched before just 'قائد' (smart filtering!).
LEGAL_KW = [
    'شكوى', 'قضية', 'محكمة', 'دعوه', 'دعوة', 'دعوى',
    'مباحث', 'الكترون',
    # military ranks - longer first
    'جندي أول', 'جندي', 'عريف', 'رقيب', 'ملازم', 'نقيب', 'رائد',
    'مقدم', 'عقيد', 'عميد', 'لواء', 'فريق أول', 'فريق', 'مشير',
    # leadership / gov - compound phrases first
    'قائد القوات', 'قائد عام', 'قائد فريق', 'قائد',
    'رئيس مجلس إدارة', 'رئيس مجلس بلدي', 'رئيس جمهورية',
    'رئيس وزراء', 'رئيس محكمة', 'رئيس تحرير', 'رئيس بلدية',
    'رئيس وحدة', 'رئيس لجنة', 'رئيس قسم', 'رئيس دولة', 'رئيس',
    'نائب رئيس الوزراء', 'نائب رئيس', 'نائب مدير', 'نائب محافظ',
    'وزير دولة', 'وزير',
    'وكيل وزارة', 'وكيل',
    'مدير عمليات', 'مدير موارد بشرية', 'مدير مشروع', 'مدير مالي',
    'مدير إداري', 'مدير تسويق', 'مدير تنفيذي', 'مدير عام', 'مدير',
    'محافظ', 'والي', 'سفير', 'قنصل',
    'أمين عام', 'سكرتير عام',
    'مستشار', 'مسؤول', 'عضو مجلس',
    'مدعي عام', 'مدعي',
    'قاضي', 'رئيس محكمة', 'مفتي', 'شيخ', 'إمام',
    # academic / professional
    'دكتور', 'أستاذ', 'مهندس', 'خبير', 'محلل',
    'مشرف', 'مراقب', 'منسق',
    'كاتب', 'صحفي', 'باحث', 'عالم', 'مخترع',
    'مؤسس', 'شريك', 'مالك',
    # royalty
    'ولي عهد', 'حاكم', 'أمير', 'ملك', 'سلطان'
]

def has_legal_kw(text):
    for w in LEGAL_KW:
        if w in text:
            return True
    return False


def try_parse_report(body_text):
    # Regex magic to catch "Report [name] : [number]". Spaces don't matter, .+? handles it!
    m = re.search(r'Report\s+(.+?)\s*:\s*(\d{8})', body_text, re.IGNORECASE)
    if m:
        return m.group(1).strip(), m.group(2)
    return None, None


def load_and_parse(mbox_path):
    try:
        engine_mailbox = mailbox.mbox(mbox_path)
    except Exception as err:
        print(f"Yikes, couldn't open the MBOX file: {err}")
        sys.exit(1)

    all_parsed_rows = []
    phone_tracker = Counter() # Huge brain way to count occurrences fast!

    print(f"Firing up the engine for {mbox_path} ... hold tight!")

    for msg in tqdm(engine_mailbox, desc="Crushing emails"):
        from_raw = msg.get('from', '') or ''
        # Plucking out just the raw email address. Regex is a lifesaver here!
        m = re.search(r'[\w.\-+]+@[\w.\-]+\.\w+', from_raw)
        pure_email = m.group(0) if m else ''

        subj = fix_numbers(msg.get('subject', '') or '')

        # Digging into the email body text. Multi-part emails are super annoying but we handle them cleanly!
        body_txt = ''
        try:
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == 'text/plain':
                        raw_bytes = part.get_payload(decode=True)
                        if raw_bytes:
                            body_txt = raw_bytes.decode('utf-8', errors='replace')
                            break  # Snag the first plain text chunk and run!
            else:
                raw_bytes = msg.get_payload(decode=True)
                if raw_bytes:
                    body_txt = raw_bytes.decode('utf-8', errors='replace')
        except Exception:
            body_txt = ''

        body_txt = fix_numbers(body_txt)
        detected_phones = pull_phones(subj + ' ' + body_txt)
        phone_tracker.update(detected_phones) # Boom, instantly count our phones

        all_parsed_rows.append({
            'sender_id': from_raw,
            'sender_email': pure_email,
            'title': subj,
            'body': body_txt,
            'phones': detected_phones
        })

    engine_mailbox.close()
    return all_parsed_rows, phone_tracker


def load_from_csv(csv_path):
    try:
        # Some CSVs might use semicolon or tab, but comma is standard. 
        # We try to be a bit flexible here.
        df = pd.read_csv(csv_path)
    except Exception as err:
        print(f"Yikes, couldn't open the CSV file: {err}")
        return [], Counter()

    all_parsed_rows = []
    phone_tracker = Counter()

    # Smart mapping: We look for columns that match our expected fields in Arabic and English!
    col_map = {
        'sender_id':    ['sender_id', 'sender', 'from', 'المرسل', 'من'],
        'sender_email': ['sender_email', 'email', 'البريد', 'بريد'],
        'title':        ['title', 'subject', 'الموضوع', 'عنوان'],
        'body':         ['body', 'content', 'text', 'المحتوى', 'النص']
    }

    def find_col(possible_names):
        for name in possible_names:
            for col in df.columns:
                if str(col).lower().strip() == name.lower():
                    return col
        return None

    mapped_cols = {k: find_col(v) for k, v in col_map.items()}

    print(f"📊 CSV columns mapped: {mapped_cols}")

    for _, row in df.iterrows():
        from_raw = str(row.get(mapped_cols['sender_id'], '')) if mapped_cols['sender_id'] else ''
        pure_email = str(row.get(mapped_cols['sender_email'], '')) if mapped_cols['sender_email'] else ''
        
        # If pure_email is missing but from_raw has an email pattern, snag it!
        if not pure_email and from_raw:
            m = re.search(r'[\w.\-+]+@[\w.\-]+\.\w+', from_raw)
            pure_email = m.group(0) if m else ''

        subj = fix_numbers(str(row.get(mapped_cols['title'], ''))) if mapped_cols['title'] else ''
        body_txt = fix_numbers(str(row.get(mapped_cols['body'], ''))) if mapped_cols['body'] else ''

        detected_phones = pull_phones(subj + ' ' + body_txt)
        phone_tracker.update(detected_phones)

        all_parsed_rows.append({
            'sender_id': from_raw,
            'sender_email': pure_email,
            'title': subj,
            'body': body_txt,
            'phones': detected_phones
        })

    return all_parsed_rows, phone_tracker


def write_outputs(outdir, all_rows, phone_tracker, spam_threshold):
    # Setting up our output folders nice and neat!
    xl_dir   = os.path.join(outdir, 'Excel_Reports')
    sql_dir  = os.path.join(outdir, 'SQL_Scripts')
    logs_dir = os.path.join(outdir, 'Summary_Logs')
    
    for folder in [xl_dir, sql_dir, logs_dir]:
        os.makedirs(folder, exist_ok=True)

    # Cool little shorthand helpers so my code below stays clean and readable
    def xl(name):  return os.path.join(xl_dir,   name)
    def sq(name):  return os.path.join(sql_dir,  name)
    def lg(name):  return os.path.join(logs_dir, name)

    # Drop the master file first. Always good to have a backup of everything.
    save_excel(all_rows, xl('all_emails.xlsx'))

    # Let's figure out who our frequent spammers are (hits >= spam_threshold)
    hot_numbers = set()
    for num, count in phone_tracker.items():
        if count >= spam_threshold:
            hot_numbers.add(num)

    # --- PIPELINE STAGES SETUP ---
    # These are my data buckets for filtering.
    stage1_spammers = []     # The worst offenders
    stage2_legal_stuff = []  # Legal / Government warnings
    stage3_reports = []      # The structured reports
    stage4_other_phones = [] # Anything else with a number
    stage5_unprocessed = []  # No phones, no legal stuff (basically useless to us)

    print("Running the magic pipeline filters...")

    for row in all_rows:
        full_text = row['title'] + ' ' + row['body']
        r_name, r_phone = try_parse_report(row['body'])

        # Filter 1: Is this one of our frequent spammer numbers?
        if any(p in hot_numbers for p in row['phones']):
            row['stage'] = 1
            stage1_spammers.append(row)
            continue

        # Filter 2: Did they use legal/sensitive keywords?
        if has_legal_kw(full_text):
            row['stage'] = 2
            stage2_legal_stuff.append(row)
            continue

        # Filter 3: Is it a structured report?
        if r_phone is not None:
            row['stage'] = 3
            row['report_name'] = r_name
            row['report_phone'] = r_phone
            stage3_reports.append(row)
            continue

        # Filter 4: Got any phones left over?
        if len(row['phones']) > 0:
            row['stage'] = 4
            stage4_other_phones.append(row)
        else:
            # Filter 5: Catch-all bin
            row['stage'] = 5
            stage5_unprocessed.append(row)

    # --- WRITING OUT STAGE 1 (Spammers) ---
    if stage1_spammers:
        # BRD says we only want phone number, count, and associated sender emails.
        mo_records = []
        for num in hot_numbers:
            count = phone_tracker[num]
            # Dig out all unique senders that had this number
            associated_senders = list(dict.fromkeys(
                r['sender_email'] for r in stage1_spammers
                if num in r['phones'] and r['sender_email']
            ))
            mo_records.append({
                'phone': num,
                'occurrences': count,
                'sender_emails': ', '.join(associated_senders)
            })
            
        # Sort by occurrences descending (highest hits on top!)
        mo_records.sort(key=lambda x: x['occurrences'], reverse=True)
        save_excel(mo_records, xl('most_occurred.xlsx'))

        # Generating SQL - straight to prod!
        with open(sq('most_occurred_block.sql'), 'w', encoding='utf-8') as f:
            for n in hot_numbers:
                f.write(f"INSERT IGNORE INTO `blocked`(`phone`, `active`) VALUES ({n},1);\n")

        unique_senders_1 = list(dict.fromkeys(r['sender_email'] for r in stage1_spammers if r['sender_email']))
        with open(lg('most_occurred_senders.txt'), 'w', encoding='utf-8') as f:
            f.write(', '.join(unique_senders_1))

    # --- WRITING OUT STAGE 2 (Legal) ---
    if stage2_legal_stuff:
        save_excel(stage2_legal_stuff, xl('legal_keywords.xlsx'))
        us2 = list(dict.fromkeys(r['sender_email'] for r in stage2_legal_stuff if r['sender_email']))
        with open(lg('legal_keywords_senders.txt'), 'w', encoding='utf-8') as f:
            f.write(', '.join(us2))

    # --- WRITING OUT STAGE 3 (Reports) ---
    if stage3_reports:
        save_excel(stage3_reports, xl('report_entries.xlsx'))
        with open(sq('report_entries_delete.sql'), 'w', encoding='utf-8') as f:
            for r in stage3_reports:
                # Gotta escape those single quotes so SQL doesn't explode!
                safe_name = str(r['report_name']).replace("'", "''")
                f.write(f"DELETE FROM `fullNames` WHERE phone = {r['report_phone']} and name = '{safe_name}' LIMIT 1;\n")
                
        us3 = list(dict.fromkeys(r['sender_email'] for r in stage3_reports if r['sender_email']))
        with open(lg('report_entries_senders.txt'), 'w', encoding='utf-8') as f:
            f.write(', '.join(us3))

    # --- WRITING OUT STAGE 4 (Other Phones) ---
    if stage4_other_phones:
        save_excel(stage4_other_phones, xl('detected_phones.xlsx'))
        
        # Grab all the unique left-over phones
        all_leftover_phones = []
        for r in stage4_other_phones:
            for p in r['phones']:
                if p not in all_leftover_phones:
                    all_leftover_phones.append(p)
                    
        with open(sq('detected_phones_block.sql'), 'w', encoding='utf-8') as f:
            for p in all_leftover_phones:
                f.write(f"INSERT IGNORE INTO `blocked`(`phone`, `active`) VALUES ({p},1);\n")
                
        us4 = list(dict.fromkeys(r['sender_email'] for r in stage4_other_phones if r['sender_email']))
        with open(lg('detected_phones_senders.txt'), 'w', encoding='utf-8') as f:
            f.write(', '.join(us4))

    # --- WRITING OUT STAGE 5 (Unprocessed Junk) ---
    save_excel(stage5_unprocessed, xl('unprocessed.xlsx'))

    # --- FINAL WRAP-UP STATS ---
    with open(lg('summary.txt'), 'w', encoding='utf-8') as f:
        f.write(f"TOTAL EMAILS DESTROYED: {len(all_rows)}\n")
        f.write(f"Stage 1 (Spammers):     {len(stage1_spammers)}\n")
        f.write(f"Stage 2 (Legal stuff):  {len(stage2_legal_stuff)}\n")
        f.write(f"Stage 3 (Reports):      {len(stage3_reports)}\n")
        f.write(f"Stage 4 (Leftovers):    {len(stage4_other_phones)}\n")
        f.write(f"Stage 5 (Useless bins): {len(stage5_unprocessed)}\n")

    print("Boom! Done. Checkout the files:")
    print(f"  Excel files -> {xl_dir}")
    print(f"  SQL Scripts -> {sql_dir}")
    print(f"  Log Stats   -> {logs_dir}")


def main():
    parser = argparse.ArgumentParser(description='Supercharged Email MBOX Processor')
    parser.add_argument('input', help='path to your .mbox file')
    parser.add_argument('--output-dir', default='./output', help='where you want me to save results')
    parser.add_argument('--spam-threshold', type=int, default=5, help='number of hits before considering it spam')
    parser.add_argument('--classify', action='store_true', help='turn on Gemini AI brains! (requires API key)')
    
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"Bruh, file not found: {args.input}")
        sys.exit(1)

    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)

    all_rows, phone_tracker = load_and_parse(args.input)

    # Let's hit the AI if the flag is on!
    if args.classify:
        print("Waking up the AI... classifying everything now.")
        batch_size = 20
        for i in range(0, len(all_rows), batch_size):
            chunk = all_rows[i : i + batch_size]
            results = run_ai_classification(chunk)
            
            if results and results[0] == "RATE_LIMIT_HIT":
                print("AI Rate Limit Hit! Continuing without AI.")
                break

            for j, category in enumerate(results):
                if i + j < len(all_rows):
                    all_rows[i + j]['category'] = category
                    
            # Tiny little sleep so we don't accidentally DDoSing the API!
            time.sleep(1)

    write_outputs(args.output_dir, all_rows, phone_tracker, args.spam_threshold)


if __name__ == '__main__':
    main()